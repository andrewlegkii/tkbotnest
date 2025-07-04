import telebot
from telebot import types
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import re

# === Настройки ===
TOKEN = '7742178744:AAHMG4v5D2IgUA82s1oYM4B1jEYPW7gzT44'
bot = telebot.TeleBot(TOKEN)

EXCEL_FILE = 'data.xlsx'
PHOTOS_FOLDER = 'photos'

# Создаем папку для фото, если её нет
os.makedirs(PHOTOS_FOLDER, exist_ok=True)

# Создаем Excel, если его нет
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Документы'
    ws.append(['№', 'Дата/время', 'ТК', 'Номер документа', 'Статус'])
    wb.save(EXCEL_FILE)

# ID администратора и пароль для выгрузки данных
admin_id = 360300829
DATA_PASSWORD = "2695"

# Хранилище состояний пользователей
user_states = {}

# === Функция загрузки информации о пользователях из users.txt ===
def load_user_companies():
    user_companies = {}
    if os.path.exists("users.txt"):
        with open("users.txt", "r", encoding="utf-8") as f:
            for line in f:
                if "-" in line:
                    parts = line.strip().split(" - ", 1)
                    if len(parts) == 2:
                        user_id_str, company = parts
                        if user_id_str.isdigit():
                            user_companies[int(user_id_str)] = company
    return user_companies


# === Команда /start ===
@bot.message_handler(commands=['start'])
def start(message):
    user_states[message.chat.id] = {}
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton("ℹ Информация"))
    markup.add(types.KeyboardButton("🆘 Помощь"))
    bot.send_message(message.chat.id, "Введите номер документа (например, R101...)", reply_markup=markup)

# === Команда /data — выгрузка файла Excel ===
@bot.message_handler(commands=['data'])
def send_data_file(message):
    args = message.text.split(maxsplit=1)
    if len(args) < 2:
        bot.reply_to(message, "❗ Пожалуйста, укажи пароль после команды, например:\n/data пароль123")
        return
    password = args[1].strip()
    if password != DATA_PASSWORD:
        bot.reply_to(message, "❌ Неверный пароль!")
        return
    try:
        with open(EXCEL_FILE, 'rb') as f:
            bot.send_document(message.chat.id, f)
    except Exception as e:
        bot.reply_to(message, f"Ошибка при отправке файла: {e}")

# === Кнопка "Информация" ===
@bot.message_handler(func=lambda message: message.text == "ℹ Информация")
def info(message):
    info_text = (
        "ℹ️ Инструкция для водителя\n\n"
        "📄 Перед загрузкой накладной на поддоны:\n\n"
        "Передайте 2 экземпляра Торг-12 на поддоны в окно приёмки документов на РЦ вместе с основными документами на груз.\n"
        "⚠️ Нельзя разделять документы на груз и поддоны — накладная на поддоны входит в комплект товарно-сопроводительной документации при поставке на РЦ Тандера.\n\n"
        "Проверьте, чтобы на Торг-12 и ТрН стояли:\n\n"
        "Подпись сотрудника РЦ\n\n"
        "Печать Тандера\n\n"
        "📸 Только после этого прикрепляйте фото документа через бот.\n\n"
    )
    bot.send_message(message.chat.id, info_text)

# === Кнопка "Назад" ===
@bot.message_handler(func=lambda message: message.text == "🔙 Назад")
def go_back(message):
    user_states[message.chat.id] = {}
    start(message)

# === Кнопка "Отправить ещё скан" ===
@bot.message_handler(func=lambda message: message.text == "📎 Отправить ещё скан")
def send_another_scan(message):
    user_states[message.chat.id] = {}
    start(message)

# === Кнопка "Помощь" ===
@bot.message_handler(func=lambda message: message.text == "🆘 Помощь")
def help_start(message):
    user_states[message.chat.id] = {'help_mode': True}
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add(types.KeyboardButton("🔙 Назад"))
    bot.send_message(message.chat.id, "Напишите сообщение администратору. Чтобы выйти, нажмите '🔙 Назад'.", reply_markup=markup)

# === Обработка текстовых сообщений ===
@bot.message_handler(func=lambda message: True)
def handle_text(message):
    state = user_states.get(message.chat.id, {})

    # Режим помощи — пересылаем админу
    if state.get('help_mode'):
        if message.text == "🔙 Назад":
            user_states[message.chat.id] = {}
            start(message)
            return
        bot.send_message(admin_id, f"Сообщение от @{message.from_user.username or message.from_user.first_name} (ID {message.chat.id}):\n{message.text}")
        bot.send_message(message.chat.id, "✅ Ваше сообщение отправлено администратору. Ожидайте ответ.")
        return

    # Обработка ввода номера документа
    if 'doc' not in state:
        doc_number = message.text.strip()

        # Проверка формата номера (R101 + цифры, максимум 10 символов)
        if not re.match(r"^[Rr]101\d{0,7}$", doc_number) or len(doc_number) > 10:
            error_msg = "❌ Неверный формат номера!\n\n" \
                      "Номер должен:\n" \
                      "• Начинаться с R101\n" \
                      "• Содержать только цифры после R101\n" \
                      "• Быть не длиннее 10 символов\n\n" \
                      "Пример: R10112345"
            bot.send_message(message.chat.id, error_msg)
            return

        user_states[message.chat.id] = {'doc': doc_number}  # Теперь ТК будет взят из users.txt
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton("🔙 Назад"))
        markup.add(types.KeyboardButton("🆘 Помощь"))
        bot.send_message(
            message.chat.id,
            "Теперь отправьте фото накладной 📷\n‼️ Убедитесь, что на Торг-12 и ТН есть:\n"
            "• Подпись сотрудника РЦ\n• Печать Тандера",
            reply_markup=markup
        )
        return

    bot.send_message(message.chat.id, "Пожалуйста, отправьте фото накладной или используйте кнопки.")

# === Обработка фото ===
@bot.message_handler(content_types=['photo'])
def handle_photo(message):
    state = user_states.get(message.chat.id)
    if not state or 'doc' not in state:
        bot.send_message(message.chat.id, "Сначала введите номер документа.")
        return

    doc_number = state['doc']
    user_id = message.chat.id
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Подгружаем данные о компаниях
    user_companies = load_user_companies()
    company_name = user_companies.get(user_id, "Неизвестная организация")

    # Сохраняем фото
    file_info = bot.get_file(message.photo[-1].file_id)
    downloaded_file = bot.download_file(file_info.file_path)

    file_path = os.path.join(PHOTOS_FOLDER, f"{doc_number}.jpg")

    with open(file_path, 'wb') as f:
        f.write(downloaded_file)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    row_number = ws.max_row + 1
    ws.append([row_number, now, company_name, doc_number, "Ожидает"])
    wb.save(EXCEL_FILE)

    caption = (
        f"📸 Новое фото от пользователя (ID: {user_id})\n"
        f"ТК: {company_name}\n"
        f"Документ: {doc_number}\n"
        f"Дата: {now}"
    )

    markup_admin = types.InlineKeyboardMarkup()
    btn_approve = types.InlineKeyboardButton("✅ Подтвердить", callback_data=f"approve_{user_id}_{doc_number}")
    btn_reject = types.InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_{user_id}_{doc_number}")
    markup_admin.add(btn_approve, btn_reject)

    with open(file_path, 'rb') as photo:
        bot.send_photo(admin_id, photo, caption=caption, reply_markup=markup_admin)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton("📎 Отправить ещё скан"))
    bot.send_message(
        message.chat.id,
        f"✅ Документ {doc_number} загружен и сохранён.\nСпасибо!",
        reply_markup=markup
    )

    user_states.pop(message.chat.id, None)

# === Обработка нажатий на кнопки подтверждения ===
@bot.callback_query_handler(func=lambda call: True)
def handle_callback_query(call):
    data = call.data.split('_')
    action = data[0]
    user_id = int(data[1])
    doc_number = data[2]

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Обновляем статус документа в Excel
    for row in ws.iter_rows(min_col=4, max_col=4):  # Ищем по столбцу D (Номер документа)
        for cell in row:
            if cell.value == doc_number:
                status_cell = ws.cell(row=cell.row, column=5)
                if action == "approve":
                    status_cell.value = "Подтверждён"
                    bot.send_message(user_id, f"✅ Документ {doc_number} успешно подтверждён!")
                    bot.answer_callback_query(call.id, "✅ Подтверждено")
                elif action == "reject":
                    status_cell.value = "Отклонён"
                    bot.send_message(user_id, f"❌ Документ {doc_number} не прошёл проверку.")
                    bot.answer_callback_query(call.id, "❌ Отклонено")
                break
    wb.save(EXCEL_FILE)

# === Команда /reply — ответ от админа пользователю (по желанию) ===
@bot.message_handler(commands=['reply'])
def admin_reply(message):
    if message.chat.id != admin_id:
        return

    parts = message.text.split(' ', 2)
    if len(parts) < 3:
        bot.send_message(admin_id, "Использование: /reply <user_id> <текст ответа>")
        return
    user_id_str, answer = parts[1], parts[2]

    if not user_id_str.isdigit():
        bot.send_message(admin_id, "ID пользователя должен быть числом.")
        return
    user_id_int = int(user_id_str)

    try:
        bot.send_message(user_id_int, f"💬 Ответ администратора:\n{answer}")
        bot.send_message(admin_id, "Ответ отправлен.")
    except Exception as e:
        bot.send_message(admin_id, f"Ошибка при отправке сообщения: {e}")

# === Запуск бота ===
if __name__ == '__main__':
    print("Bot is running...")
    bot.infinity_polling(timeout=60, long_polling_timeout=60)
